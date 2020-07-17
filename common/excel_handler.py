#!/usr/bin/env python
# -*- coding:utf-8 -*-

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import statistics


# @data list 格式数据
def write_list_2_excel(filename, data):
    wb = Workbook()
    sheet = wb.active
    for row in data:
        sheet.append(row)
    wb.save(filename)


def read_excel(filename):
    # 默认可读写，若有需要可以指定write_only和read_only为True
    # step 1：打开文件（定只读模式, 性能更好）[data_only=True 如果单元格式公式，将输出计算结果 eg：=AVERAGE(B2:B8)]
    wb = load_workbook(filename, read_only=True, data_only=True)
    wb = load_workbook(filename)

    # step 2：获取sheet
    # 2.1 获取所有sheet名称
    sheet_names = wb.sheetnames

    # 2.2 根据名字获取sheet
    sheet_1 = wb['OneSheet']
    sheet_2 = wb['Sheet2']
    sheet_3 = wb['Sheet3']

    # 2.3 获取当前激活的sheet
    sheet = wb.active

    # step 3 : 获取单元格
    # 方法一： 用sheet下标获得
    b4 = sheet['B4']
    # 分别返回
    print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型

    # 方法二： 用cell函数, 换成数字
    c4 = sheet.cell(row=4, column=3)
    print(c4.value)           # c4 的值
    print(c4.coordinate)      # c4 的坐标

    # 3.2 获得最大行和最大列
    print(sheet.max_row)
    print(sheet.max_column)

    # 3.3 获取行和列
    # 3.3.1 按行，所以返回A1, B1, C1这样的顺序
    for row in sheet.rows:
        for cell in row:
            try:
                print(cell.coordinate)
            except AttributeError:
                pass

    print('-' * 80)

    # 3.3.2 按列，所以返回A1, A2, A3 这样的顺序
    for column in sheet.columns:
        for cell in column:
            try:
                print(cell.coordinate)
            except AttributeError:
                pass

    print('-' * 80)

    # 3.4 要获得某行的数据 使用索引  ps：sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
    for cell in list(sheet.rows)[2]:
        print(cell.value)

    print('-' * 80)

    # 3.5 要获得任意区间单元格 eg：A1-B3
    for row in range(1, 4):
        for column in range(1, 3):
            print(sheet.cell(row=row, column=column))

    print('-' * 80)

    # 3.6 使用切片 eg：sheet['A1':'B3']
    for row_cell in sheet['A1':'B3']:
        for cell in row_cell:
            print(cell)

    print('*' * 80)

    # 3.6 使用切片 eg：sheet['A1':'B3']
    for cell in sheet['A1':'B3']:
        print(cell)

    # 读取多行
    cells = sheet['A1': 'B6']
    for c1, c2 in cells:
        print("{0:8} {1:8}".format(c1.value, c2.value))

    print('-' * 80)

    # 番外：字母获得列号，根据列号返回字母
    # 根据列的数字返回字母
    print(get_column_letter(2))
    # 根据字母返回数字
    print(column_index_from_string('B'))


def write_excel(filename):
    # step 1: 导入 WorkBook （只是新建，但未保存）
    #         指定只写模式, 性能更好
    # wb = Workbook(write_only=True)
    wb = Workbook()

    # step 2: 创建 sheet 默认第1个位置 index=0
    sheet = wb.create_sheet(title='FirstSheetNew', index=0)
    sheet2 = wb.create_sheet(title='SecondSheetNew', index=1)
    print(wb.sheetnames)

    # 删除某个工作表
    wb.remove(sheet2)

    # step 3: 写入单元格
    # 3.1 直接给单元格赋值就行
    sheet['A1'] = 'good'
    # 3.2 B9处写入平均值
    sheet['B9'] = '=AVERAGE(B2:B8)'
    sheet['A2'] = 1
    sheet['A3'] = 'world'

    # step 4：一次写入多行数据 ：append函数
    # 4.1 添加1行
    row = [1, 2, 3, 4, 5]
    sheet.append(row)
    # 添加多行
    rows = [
        ['Number', 'data1', 'data2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
    ]
    for row in rows:
        sheet.append(row)

    # 添加多列，将多行列表，转置成元祖，再形成列表
    columns = list(zip(*rows))
    for col in columns:
        sheet.append(col)

    # step 5: 设置字体， 对齐方式， 设置行高和列宽, 合并和拆分单元格
    # 设置字体
    bold_italic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
    sheet['A1'].font = bold_italic_24_font

    # 对齐方式
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置 第2行 行高
    sheet.row_dimensions[1].height = 60
    # 设置 C列 列宽
    sheet.column_dimensions['C'].width = 60

    # 合并单元格， 往左上角写入数据即可
    sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
    sheet.merge_cells('A3:C5')  # 合并一个矩形区域中的单元格

    # 拆分单元格
    sheet.unmerge_cells('A1:C3')

    wb.save(filename)


def update_excel(filename):
    wb = load_workbook(filename)
    sheet = wb.active

    sheet['B1'] = 100000
    sheet.cell(row=2, column=2, value=20000000)
    sheet.cell(row=2, column=3).value = 30000000

    rows = (
        (88, 46, 57),
        (89, 38, 12),
        (23, 59, 78),
        (56, 21, 98),
        (24, 18, 43),
        (34, 15, 67)
    )

    for row in rows:
        sheet.append(row)

    # iter_rows() 方法将工作表中的单元格返回为行
    for row in sheet.iter_rows(min_row=21, min_col=1, max_row=26, max_col=3):
        for cell in row:
            print(cell.value, end=" ")
        print()

    # iter_cols() 方法将工作表中的单元格作为列返回
    for col in sheet.iter_cols(min_row=21, min_col=1, max_row=26, max_col=3):
        for cell in col:
            print(cell.value, end=" ")
        print()

    wb.save(filename)


# 统计
def stats(filename):
    wb = load_workbook(filename, data_only=True)
    sheet = wb.active
    values = []

    for row in sheet.rows:
        for cell in row:
            values.append(cell.value)

    print("Number of values: {0}".format(len(values)))
    print("Sum of values: {0}".format(sum(values)))
    print("Minimum value: {0}".format(min(values)))
    print("Maximum value: {0}".format(max(values)))
    print("Mean 平均值 : {0}".format(statistics.mean(values)))
    print("Median 中位数: {0}".format(statistics.median(values)))
    print("Standard deviation 标准偏差 : {0}".format(statistics.stdev(values)))
    print("Variance 方差: {0}".format(statistics.variance(values)))


# 过滤 和 排序
def filter_sort(filename):
    wb = Workbook()
    sheet = wb.active

    data = [
        ['Item', 'Colour'],
        ['pen', 'brown'],
        ['book', 'black'],
        ['plate', 'white'],
        ['chair', 'brown'],
        ['coin', 'gold'],
        ['bed', 'brown'],
        ['notebook', 'white'],
    ]

    for r in data:
        sheet.append(r)

    sheet.auto_filter.ref = 'A1:B8'
    sheet.auto_filter.add_filter_column(1, ['brown', 'white'])
    sheet.auto_filter.add_sort_condition('B2:B8')

    wb.save(filename)


